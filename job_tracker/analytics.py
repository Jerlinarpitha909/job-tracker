"""
Analytics: query statistics and export to CSV / Excel.
"""

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from job_tracker.models import JobRepository, JobApplication
from job_tracker.automation import days_since_applied

EXPORTS_DIR = Path(__file__).parent.parent / "exports"


def get_analytics(repo: JobRepository = None) -> dict:
    """
    Returns a dict of key metrics:
      - total_applied
      - by_status      (counts per status)
      - response_rate  (% that moved past 'Applied')
      - avg_days_to_response  (avg days for jobs that progressed)
      - oldest_open    (days for longest-open application)
    """
    repo = repo or JobRepository()
    jobs = repo.get_all()

    if not jobs:
        return {
            "total_applied": 0,
            "by_status": {},
            "response_rate": 0.0,
            "avg_days_to_response": 0.0,
            "oldest_open": 0,
        }

    by_status: dict[str, int] = {}
    for job in jobs:
        by_status[job.status] = by_status.get(job.status, 0) + 1

    progressed = [j for j in jobs if j.status not in ("Applied", "Follow Up Needed")]
    response_rate = (len(progressed) / len(jobs)) * 100 if jobs else 0.0

    days_list = [days_since_applied(j) for j in progressed if days_since_applied(j) >= 0]
    avg_days = sum(days_list) / len(days_list) if days_list else 0.0

    open_jobs = [j for j in jobs if j.status in ("Applied", "Interview", "Follow Up Needed")]
    oldest = max((days_since_applied(j) for j in open_jobs), default=0)

    return {
        "total_applied": len(jobs),
        "by_status": by_status,
        "response_rate": round(response_rate, 1),
        "avg_days_to_response": round(avg_days, 1),
        "oldest_open": oldest,
    }


def export_to_csv(repo: JobRepository = None, filepath: Optional[Path] = None) -> Path:
    """Exports all applications to a CSV file. Returns the file path."""
    repo = repo or JobRepository()
    jobs = repo.get_all()
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filepath = filepath or EXPORTS_DIR / f"applications_{date.today().isoformat()}.csv"

    headers = ["id", "company", "role", "status", "date_applied", "notes", "last_updated"]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for job in jobs:
            writer.writerow({
                "id": job.id,
                "company": job.company,
                "role": job.role,
                "status": job.status,
                "date_applied": job.date_applied,
                "notes": job.notes,
                "last_updated": job.last_updated,
            })
    return filepath


def export_to_excel(repo: JobRepository = None, filepath: Optional[Path] = None) -> Path:
    """Exports all applications to an Excel (.xlsx) file. Returns the file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    repo = repo or JobRepository()
    jobs = repo.get_all()
    analytics = get_analytics(repo)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filepath = filepath or EXPORTS_DIR / f"applications_{date.today().isoformat()}.xlsx"

    wb = openpyxl.Workbook()

    # --- Sheet 1: Applications ---
    ws = wb.active
    ws.title = "Applications"
    headers = ["ID", "Company", "Role", "Status", "Date Applied", "Notes", "Last Updated"]
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        "Applied":          "BDD7EE",
        "Interview":        "FFF2CC",
        "Offer":            "E2EFDA",
        "Rejected":         "FCE4D6",
        "Follow Up Needed": "F4B8C1",
    }

    for row_idx, job in enumerate(jobs, 2):
        row_data = [job.id, job.company, job.role, job.status,
                    job.date_applied, job.notes, job.last_updated]
        fill_color = status_colors.get(job.status, "FFFFFF")
        fill = PatternFill("solid", start_color=fill_color)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    col_widths = [6, 22, 30, 20, 14, 40, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # --- Sheet 2: Analytics ---
    ws2 = wb.create_sheet("Analytics")
    ws2["A1"] = "Job Application Analytics"
    ws2["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Total Applied", analytics["total_applied"]),
        ("Response Rate (%)", analytics["response_rate"]),
        ("Avg Days to Response", analytics["avg_days_to_response"]),
        ("Oldest Open Application (days)", analytics["oldest_open"]),
    ]
    for i, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=val)

    ws2["A8"] = "Applications by Status"
    ws2["A8"].font = Font(bold=True)
    for i, (status, count) in enumerate(analytics["by_status"].items(), 9):
        ws2.cell(row=i, column=1, value=status)
        ws2.cell(row=i, column=2, value=count)

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16

    wb.save(filepath)
    return filepath
